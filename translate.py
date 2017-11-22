#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# author: Alexey Koshevoy

from openpyxl import load_workbook, Workbook, worksheet
import re


class QuTable:

    def __init__(self, path, sheet):
        self.path = path
        self.sheet_name = sheet
        self.sheet = self.get_sheet()
        self.wb = self.create_new_excel()

    @staticmethod
    def read_table(path):
        wb = load_workbook(path, read_only=True)
        return wb

    def get_sheet(self):
        """
        Make it optional – only if there is some sheets??

        Maybe we can use this:

        # >>> df = DataFrame(ws.values)
        """
        wb = self.read_table(path=self.path)
        sheet = wb[self.sheet_name]
        return sheet

    def get_horizontal_list(self):

        horizontal_list = []

        sheet = self.sheet
        for row in range(2, sheet.max_row):
            for column in "A":
                cell_name = "{}{}".format(column, row)
                horizontal_list.append(sheet[cell_name].value)

        return horizontal_list

    @staticmethod
    def iter_rows(ws):
        for row in ws.iter_rows():
            yield [cell.value for cell in row]

    def get_vertical_list(self):
        sheet = self.sheet
        vertical_list = list(self.iter_rows(sheet))[0]
        return vertical_list

    @staticmethod
    def create_new_excel():
        wb = Workbook()
        return wb

    def write(self, hor_list, ver_list):
        wb = self.wb
        ws1 = wb.create_sheet("{}_t".format(self.sheet_name))

        if hor_list:
            for i in range(len(hor_list)):
                ws1(row=1,col=i).value = hor_list[i]

        if ver_list:
            for i in range(len(ver_list)):
                ws1(row=i, col=1).value = ver_list[i]

        wb.save(filename='I_tried.xlsx')


class Translation:

    def __init__(self, list_of_words):
        self.list_of_words = list_of_words

    @staticmethod
    def get_dictionary():
        path = \
            '/Users/alexey/Documents/Python/automated_lt/dictionary/ru-de.txt'
        dictionary = open(path).readlines()
        return dictionary

    @staticmethod
    def translate_word(word, dictionary):
        for line in dictionary:
            if re.match(word, line):
                return re.search('([A-Z]|[a-z])+', line).group(0)

    def change_list(self, dictionary):

        list_words = self.list_of_words

        for i in range(len(list_words)):
            if re.match('([А-Я]?)[а-я]* (.*)', list_words[i]):
                word1 = self.translate_word(re.search('(([А-Я]?)[а-я]*) (.*)',
                                                      list_words[i]).group(1),
                                            dictionary=dictionary)
                word2 = self.translate_word(re.search('([А-Я]?)[а-я]* ((.*))',
                                                      list_words[i]).group(1),
                                            dictionary=dictionary)
                list_words[i] = '{} ({})'.format(word1, word2)

            list_words[i] = self.\
                translate_word(list_words[i], dictionary=dictionary)
        return list_words


a = QuTable\
    (path='/Users/alexey/Documents/Python/automated_lt/questionnaire_size.xlsx'
     , sheet='русский_стандртный_вид')
list_w = a.get_horizontal_list()
list_v = a.get_vertical_list()

b = Translation(list_w)
# s = Translation(list_v)
list_w_t = b.change_list(b.get_dictionary())
# list_v_t = s.change_list(b.get_dictionary())

print(list_w_t)
#
# a.write(list_w_t, list_v_t)
